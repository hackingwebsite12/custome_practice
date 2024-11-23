import pytest
import openpyxl

def getRowcount(file,sheetname):
    workbook = openpyxl.load_workbook(file) # it will load the excel workbook in the given file#
    sheet = workbook[sheetname] # access a specific worksheet
    return (sheet.max_row) # get the maximum row number containing data

def getcolumncount(file,sheetname):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    return (sheet.max_column)

def readdata(file,sheetname,rownum,columnno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    return sheet.cell(row=rownum,column=columnno).value # get the value of a specific cell

def writedata(file,sheetname,rownum,columnno,data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    sheet.cell(row=rownum,column=columnno).value = data # set the values of a specific cell
    workbook.save(file)